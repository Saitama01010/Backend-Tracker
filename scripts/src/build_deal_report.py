import json, glob, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

deals = json.load(open("/tmp/deals.json"))
results = json.load(open("/tmp/results.json"))

headers = [
    "Agent Name", "Agent Team", "Customer Name", "File ID", "Phone", "Submission Date", "Status",
    "Total Calls", "Retention (in)", "NSF", "CS", "Onboarding", "OB (outbound)", "Other (in)",
    "Live Call?", "Transferred From (Company)", "Transferred By (Agent)", "Live Call Evidence",
    "Transfer Source", "Flag", "Outcome Summary",
]
COL = {h: i + 1 for i, h in enumerate(headers)}

widths = {
    "Agent Name": 22, "Agent Team": 14, "Customer Name": 18, "File ID": 12, "Phone": 14,
    "Submission Date": 18, "Status": 22, "Total Calls": 9, "Retention (in)": 11, "NSF": 7,
    "CS": 7, "Onboarding": 11, "OB (outbound)": 12, "Other (in)": 9, "Live Call?": 9,
    "Transferred From (Company)": 16, "Transferred By (Agent)": 18, "Live Call Evidence": 38,
    "Transfer Source": 18, "Flag": 30, "Outcome Summary": 60,
}
center_cols = {"File ID", "Total Calls", "Retention (in)", "NSF", "CS", "Onboarding",
               "OB (outbound)", "Other (in)", "Live Call?", "Transferred From (Company)"}

wb = Workbook()
ws = wb.active
ws.title = "May Deals - Call Analysis"

hdr_fill = PatternFill("solid", fgColor="3B0764")
hdr_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D4D4D8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="top")
wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

ws.append(headers)
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border

yes_fill = PatternFill("solid", fgColor="DCFCE7")
flag_fill = PatternFill("solid", fgColor="FEF9C3")
nocall_fill = PatternFill("solid", fgColor="FEE2E2")

for d in deals:
    r = results.get(d.get("_e164") or "", {})
    total = r.get("total_calls", 0)
    if total == 0:
        flag = "No calls found"
    elif r.get("ob_done_no_retention"):
        flag = "OB done — no calls on retention line"
    else:
        flag = ""
    row = [
        d.get("AgentName"), d.get("AgentTeam"), d.get("CustomerName"), d.get("FileID"),
        d.get("Phone"), d.get("SubmissionDate"), d.get("Status"),
        total, r.get("retention_in_completed", 0), r.get("nsf_completed", 0),
        r.get("cs_completed", 0), r.get("onboarding_completed", 0),
        r.get("ob_completed", 0), r.get("other_completed", 0),
        r.get("live_call", "No"), r.get("transfer_company", ""), r.get("transfer_agent", ""),
        r.get("live_call_evidence", ""), r.get("transfer_source", ""), flag,
        r.get("outcome_summary", ""),
    ]
    ws.append(row)
    ri = ws.max_row
    for h, ci in COL.items():
        cell = ws.cell(row=ri, column=ci)
        cell.border = border
        cell.alignment = center if h in center_cols else wrap
    if r.get("live_call") == "Yes":
        c = ws.cell(row=ri, column=COL["Live Call?"]); c.fill = yes_fill; c.font = Font(bold=True, color="166534")
    fc = ws.cell(row=ri, column=COL["Flag"])
    if flag == "No calls found":
        fc.fill = nocall_fill
    elif flag:
        fc.fill = flag_fill

for h, w in widths.items():
    ws.column_dimensions[get_column_letter(COL[h])].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

src = sorted(f for f in glob.glob("attached_assets/May_Deals_*.xlsx") if "_Call_Analysis" not in f)[-1]
base = os.path.splitext(os.path.basename(src))[0]
out = f"attached_assets/{base}_Call_Analysis.xlsx"
wb.save(out)
print("wrote", out, "rows:", ws.max_row - 1)

live = sum(1 for d in deals if results.get(d.get("_e164") or "", {}).get("live_call") == "Yes")
nocall = sum(1 for d in deals if results.get(d.get("_e164") or "", {}).get("total_calls", 0) == 0)
obdone = sum(1 for d in deals if results.get(d.get("_e164") or "", {}).get("ob_done_no_retention"))
comp = {}
for d in deals:
    r = results.get(d.get("_e164") or "", {})
    if r.get("live_call") == "Yes":
        comp[r.get("transfer_company") or "(unspecified)"] = comp.get(r.get("transfer_company") or "(unspecified)", 0) + 1
print(f"live={live} no_calls={nocall} ob_done={obdone} | by company: {comp}")
