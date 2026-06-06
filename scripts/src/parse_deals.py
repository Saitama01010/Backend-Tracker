import json, glob, re
from openpyxl import load_workbook

src = sorted(f for f in glob.glob("attached_assets/May_Deals_*.xlsx") if "_Call_Analysis" not in f)[-1]
wb = load_workbook(src, data_only=True)
ws = wb.active

cols = ["AgentName", "AgentTeam", "SplitWith", "CustomerName", "FileID",
        "DebtLoad", "SubmissionDate", "Stage", "Status", "Phone"]


def e164(raw):
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) == 10:
        return "+1" + digits
    if len(digits) == 11 and digits.startswith("1"):
        return "+" + digits
    return None


deals = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row is None or all(c is None for c in row):
        continue
    d = {cols[i]: (row[i] if i < len(row) else None) for i in range(len(cols))}
    if d.get("SubmissionDate") is not None:
        d["SubmissionDate"] = str(d["SubmissionDate"])
    d["_e164"] = e164(d.get("Phone"))
    deals.append(d)

json.dump(deals, open("/tmp/deals.json", "w"), default=str)
phones = {d["_e164"] for d in deals if d["_e164"]}
print(f"src={src} rows={len(deals)} unique_phones={len(phones)}")
